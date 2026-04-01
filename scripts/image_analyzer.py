import os
import io
import json
import base64
from PIL import Image
import openpyxl
from openpyxl.drawing.image import Image as OpenPyXLImage

class VisualAnalyzer:
    """
    [Phase 2] Visual Reference Analyzer
    Extracts images from Excel and uses Gemini Vision to analyze Visual DNA.
    """
    def __init__(self, api_key):
        self.api_key = api_key
        
    def extract_images_from_excel(self, excel_path, target_sheet_names=None):
        """
        Extracts images from specified sheets in an Excel file.
        Returns a list of PIL Image objects.
        """
        if not os.path.exists(excel_path):
            print(f"Error: Excel file not found at {excel_path}")
            return []

        images = []
        try:
            wb = openpyxl.load_workbook(excel_path, data_only=True)
            
            # Determine sheets to scan
            sheets_to_scan = []
            if target_sheet_names:
                sheets_to_scan = [s for s in wb.sheetnames if any(t in s for t in target_sheet_names)]
            else:
                # Default interesting sheets
                sheets_to_scan = [s for s in wb.sheetnames if any(k in s for k in ["風格", "Style", "Reference", "參考", "Symbol", "設計"])]
            
            print(f"Scanning sheets for images: {sheets_to_scan}")

            for sheet_name in sheets_to_scan:
                ws = wb[sheet_name]
                # openpyxl stores images in ws._images
                if hasattr(ws, '_images'):
                    for img in ws._images:
                        # Extract image data
                        img_data = img._data()
                        image = Image.open(io.BytesIO(img_data))
                        images.append(image)
                        print(f"  Found image in {sheet_name}, size: {image.size}")
                        
                        if len(images) >= 5: # Limit to 5 reference images to save tokens/time
                            break
                if len(images) >= 5:
                    break
                    
            wb.close()
            
        except Exception as e:
            print(f"Error extracting images: {e}")
            
        print(f"Total extracted images: {len(images)}")
        return images

    def analyze_visual_dna(self, images):
        """
        Sends images to Gemini Vision to extract Visual DNA (Color, Material, Lighting).
        Returns a JSON-compatible dictionary.
        """
        if not images:
            return None
            
        try:
            from google import genai
            from google.genai import types
            client = genai.Client(api_key=self.api_key)

            # Convert images to bytes for API
            image_contents = []
            for img in images:
                # Resize if too large to save bandwidth
                if img.width > 1024 or img.height > 1024:
                    img.thumbnail((1024, 1024))
                
                buf = io.BytesIO()
                img.save(buf, format=img.format if img.format else 'PNG')
                image_contents.append(types.Part.from_bytes(data=buf.getvalue(), mime_type="image/png"))

            prompt = """
            You are an Elite Art Director with expertise in visual analysis. Analyze these reference images from a slot game design document.
            Extract the comprehensive "Visual DNA" into a structured JSON format.
            
            **CRITICAL**: You MUST identify AT LEAST 3 characteristics in EACH category below. Be thorough and specific.
            **[Deep Reasoning Enabled]**: Use Advanced Chain of Thought analysis to cross-reference multiple visual elements and extract deeper thematic meanings and nuanced style traits that are not immediately obvious.
            
            Analyze these dimensions:
            
            1.  **Color Tone**: Identify the overall color temperature and mood
                - Warm (reds, oranges, yellows) / Cool (blues, greens, purples) / Neutral (grays, browns)
                - Examples: "Warm golden tones", "Cool cyberpunk blues", "Neutral earth tones"
            
            2.  **Color Palette**: Dominant hex codes and color harmony
                - List 3-5 dominant colors with hex codes
                - Describe color relationships (e.g., "High Contrast Neon", "Warm Gold & Red", "Pastel Rainbow")
            
            3.  **Materials & Textures**: Identify at least 3 material types
                - Examples: "Brushed Gold Metal", "Glossy Gemstones", "Matte Stone", "Translucent Glass", 
                  "Soft Fabric", "Rough Wood", "Liquid/Jelly", "Crystalline", "Metallic Reflective"
                - Describe surface properties: reflective, matte, rough, smooth, translucent
            
            4.  **Lighting Style**: Describe lighting setup and characteristics
                - Type: Backlit, Rim lighting, Soft studio, Hard directional, Ambient, Dramatic
                - Quality: Strong highlights, Soft shadows, High contrast, Flat/even
                - Examples: "Backlit rim lighting with strong highlights", "Soft diffused studio lighting"
            
            5.  **Outline & Edges**: Describe edge treatment
                - Thick outlines / Thin outlines / No outlines / Glowing edges
                - Examples: "Thick black cartoon outlines", "Thin cel-shaded edges", "No outlines (realistic)"
            
            6.  **Saturation & Contrast**: Describe color intensity and value range
                - Saturation: High (vibrant) / Medium / Low (desaturated)
                - Contrast: High (strong blacks & whites) / Medium / Low (soft, muted)
                - Examples: "High saturation with strong contrast", "Medium saturation with soft contrast"
            
            7.  **Style Keywords**: Overall artistic style (at least 3 keywords)
                - Examples: "Q-Version 3D", "Realistic PBR", "2D Vector", "Anime Cel-Shaded", 
                  "Painterly", "Photorealistic", "Stylized 3D", "Flat Design", "Hand-Painted"
                  
            8.  **Thematic Reasoning (NEW)**: Extract deep thematic connections
                - Examples: "Ancient Mysticism", "Futuristic Rebellion", "Whimsical Nature"
            
            Return ONLY this JSON structure (ensure ALL fields have at least 3 items/characteristics):
            {
                "color_tone": "string (warm/cool/neutral with description)",
                "palette": ["#RRGGBB", "#RRGGBB", "#RRGGBB"],
                "color_desc": "string (color harmony description)",
                "materials": ["material 1", "material 2", "material 3"],
                "lighting": "string (detailed lighting description)",
                "outline": "string (edge treatment description)",
                "saturation": "string (high/medium/low with description)",
                "contrast": "string (high/medium/low with description)",
                "style_keywords": ["keyword 1", "keyword 2", "keyword 3"],
                "thematic_reasoning": ["theme 1", "theme 2", "theme 3"]
            }
            """
            
            response = client.models.generate_content(
                model='gemini-2.5-pro',
                contents=[prompt, *image_contents]
            )
            
            # Clean up response to get JSON
            text = response.text.strip()
            if text.startswith("```json"):
                text = text[7:-3]
            elif text.startswith("```"):
                text = text[3:-3]
                
            return json.loads(text)

        except Exception as e:
            print(f"Error analyzing visual visual DNA: {e}")
            return None

# Integration Helper
def analyze_excel_style(excel_path, api_key):
    analyzer = VisualAnalyzer(api_key)
    print(f"Analyzing images in {excel_path}...")
    images = analyzer.extract_images_from_excel(excel_path)
    if not images:
        print("No reference images found.")
        return None
        
    dna = analyzer.analyze_visual_dna(images)
    print("Visual DNA Visualization Result:", json.dumps(dna, indent=2, ensure_ascii=False))
    return dna

if __name__ == "__main__":
    # Load .env
    try:
        from dotenv import load_dotenv
        load_dotenv()
        print("Loaded environment variables.")
    except ImportError:
        print("Warning: python-dotenv not installed, assuming env vars are set.")

    # Test run
    key = os.environ.get("GEMINI_API_KEY")
    if key:
        # Find an Excel file to test
        files = [f for f in os.listdir(".") if f.endswith(".xlsx")]
        if files:
            # Prefer KnockoutClash if available
            target = next((f for f in files if "KnockoutClash" in f), files[0])
            analyze_excel_style(target, key)
        else:
            print("No Excel file found for testing.")
    else:
        print("GEMINI_API_KEY not set.")
