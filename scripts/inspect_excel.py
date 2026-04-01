
import pandas as pd
import sys
import openpyxl
from openpyxl.utils import get_column_letter

if sys.stdout.encoding != 'utf-8':
    import io
    sys.stdout = io.TextIOWrapper(sys.stdout.buffer, encoding='utf-8')

def search_keywords_openpyxl(file_path, keywords):
    print(f"Loading '{file_path}' using openpyxl...")
    try:
        wb = openpyxl.load_workbook(file_path, data_only=True)
    except Exception as e:
        print(f"Error loading Excel: {e}")
        return

    found = False
    for sheet_name in wb.sheetnames:
        print(f"Scanning sheet '{sheet_name}'...")
        ws = wb[sheet_name]
        
        # Check Sheet Name
        sheet_match = False
        for kw in keywords:
            if kw in sheet_name:
                print(f"[FOUND SHEET] Name: '{sheet_name}' matches keyword '{kw}'")
                sheet_match = True
                break
        
        if sheet_match:
            print(f"--- Dense Snippet of '{sheet_name}' ---")
            for r in range(1, 100):
                row_cells = [ws.cell(row=r, column=c).value for c in range(1, 15)]
                non_empty = [str(v) for v in row_cells if v is not None and str(v).strip() != ""]
                if len(non_empty) >= 2: # Lower threshold for sheet match
                    print(f"Row {r}: {' | '.join([str(v) for v in row_cells])}")
            print("---------------------------------")
            found = True
            continue # Already dumped, skip cell-by-cell keyword search for this sheet

        for row in ws.iter_rows():
            for cell in row:
                cell_val = str(cell.value) if cell.value is not None else ""
                for kw in keywords:
                        if kw in cell_val:
                            print(f"[FOUND] Sheet: '{sheet_name}', Row: {cell.row}, Col: {cell.column}, Value: '{cell_val}'")
                            
                            # Dump sheet snippet
                            print(f"--- Dense Snippet of '{sheet_name}' ---")
                            for r in range(1, 100): # Scan up to 100 rows
                                row_cells = [ws.cell(row=r, column=c).value for c in range(1, 15)]
                                non_empty = [str(v) for v in row_cells if v is not None and str(v).strip() != ""]
                                if len(non_empty) >= 3:
                                    print(f"Row {r}: {' | '.join([str(v) for v in row_cells])}")
                            print("---------------------------------")
                            
                        found = True
                        
    if not found:
        print("No keywords found.")

if __name__ == "__main__":
    if len(sys.argv) < 2:
        print("Usage: python inspect_excel.py <excel_path>")
    else:
        path = sys.argv[1]
        search_keywords_openpyxl(path, ["Symbol設計", "Symbol設計", "Symbol說明"])
